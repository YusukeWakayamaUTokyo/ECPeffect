# coding: UTF-8
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
import time
import math
from pymeasure.instruments.keithley import Keithley2400
from datetime import datetime
import concurrent.futures
import requests # for slack
import numpy as np
import os

### thermistor settings ###
pulse_interval = 0.4
pulse_current = 0.000001 # 1uA

#thermistor 104JT-025
B_constant = 4390
R25 = 100000
T25 = 298.15

#You can change GPIB as you like
keithley1 = Keithley2400("GPIB::22") # thermistor 
keithley2 = Keithley2400("GPIB::24") # current_apply

voltage_interval = 0.2

def initial_settings():
    sheet.cell(row = 1, column = 1).value = 'time (s)'
    sheet.cell(row = 1, column = 2).value = 'Temperature (℃)'
    sheet.cell(row = 1, column = 4).value = 'Thermistor'
    sheet.cell(row = 1, column = 12).value = 'time (s)'
    sheet.cell(row = 1, column = 13).value = 'Voltage (V)'
    sheet.cell(row = 2, column = 4).value = 'pulse_current (μA)' #current value in μA
    sheet.cell(row = 2, column = 5).value = pulse_current*1000000 #current value in μA
    sheet.cell(row = 3, column = 4).value = 'pulse_interval (s)'
    sheet.cell(row = 3, column = 5).value = pulse_interval
    sheet.cell(row = 5, column = 4).value = 'Current_apply'
    sheet.cell(row = 6, column = 4).value = 'current (mA)'
    sheet.cell(row = 7, column = 4).value = 'period_number+1'
    sheet.cell(row = 8, column = 4).value = 'half_period_time'
    sheet.cell(row = 10, column = 4).value = 'Voltage_recording'
    sheet.cell(row = 11, column = 4).value = 'recording_interval (s)'
    sheet.cell(row = 1, column = 3).value = datetime.now()

    ### keithley settings ###
    keithley1.reset()
    keithley1.disable_buffer()
    keithley1.use_front_terminals()
    keithley1.apply_current()
    keithley1.source_voltage_range = 0.1
    keithley1.source_current = pulse_current
    keithley1.enable_source()
    keithley1.measure_voltage()
    keithley1.compliance_voltage = 1.8

    keithley2.reset()    
    keithley2.disable_buffer()
    keithley2.use_front_terminals()
    keithley2.apply_current()
    keithley2.source_voltage_range = 0.3   
    keithley2.source_current = 0
    keithley2.enable_source()
    keithley2.measure_voltage()
    keithley2.compliance_voltage = 1.3

# print the estimated time (h/m/s) for measurement before starting
def time_calculation():
    # in the unit of second
    total_required_time = before_measurement_h*3600 + measurement_interval_h*3600 * (measurements-1) + measurements   
    for j in range(measurements) :
        total_required_time = total_required_time + (period_number[j]+1) * half_period_time[j] *2  
    
    # convert into (h/m/s)
    required_time_h = int(total_required_time/3600)
    required_time_m = int((total_required_time-required_time_h*3600)/60)
    required_time_s = int(total_required_time -(required_time_h*3600+required_time_m*60))
    print('total required_time: ', required_time_h,' h ',required_time_m,' m ',required_time_s,' s ')

def ec_confirmation():
    print('sample name:', sample_name)
    print('\n')
    print('before measurement:', before_measurement_h,'h')
    for i in range(measurements):
        print('experiment', i+1)
        print('    ',I[i],'mA,',half_period_time[i],'s,',period_number[i],'+1 periods')
    print('measurement interval:',measurement_interval_h,'h')

def connection_test():
    keithley1.reset()
    keithley1.disable_buffer()
    keithley1.use_front_terminals()
    keithley1.apply_current()
    keithley1.source_current = 0
    keithley1.enable_source()

    keithley2.reset()
    keithley2.disable_buffer()
    keithley2.use_front_terminals()
    keithley2.apply_current()
    keithley2.source_current = 0
    keithley2.enable_source()

    time.sleep(0.5)
    keithley1.shutdown()
    keithley2.shutdown()    

    book = openpyxl.Workbook() # create an excel file
    sheet = book.worksheets[0] 
    print('connection OK')
    
# apply current to a thermistor and get a voltage value
def thermistor():
    target_time = 0
    time_accuracy = 0.001 # time for sleep
    rest_until_target_time = time.time() - base_time - target_time
    for step in range(thermistor_steps):
        while rest_until_target_time < 0: # wait until reaching target time
            time.sleep(time_accuracy) #wait for time_accuracy, then recalculate target time
            rest_until_target_time = time.time() - base_time - target_time
        keithley1.source_current = pulse_current # when reaching the target time, apply current
        keithley1.ramp_to_current(pulse_current,steps=1, pause = 0.000001)
        t2 = time.time()
        V = keithley1.voltage
        keithley1.source_current = 0 # after obtaining data turn off the current immediately

        R = V/pulse_current
        temperature = 1/(1/T25 + math.log(R/R25)/B_constant) - 273.15
        sheet.cell(row = step + 2, column = 1).value = t2 - base_time
        sheet.cell(row = step + 2, column = 2).value = temperature # calculate a temperature from a voltage value, and record to the excel file

        target_time = target_time + pulse_interval
        rest_until_target_time = time.time() - base_time - target_time #reset the target time
    keithley1.shutdown()
    print('thermistor_done.')

# apply a current (alternating square wave current) to the ECP cell 
def current_apply(current):
    target_time = half_period_time[i]
    voltage_time = voltage_interval
    time_accuracy = 0.001

    rest_until_target_time = time.time() - base_time - target_time
    for step in range((period_number[i]+1) *2):
        keithley2.ramp_to_current(current,steps=3)
        while rest_until_target_time < 0: # wait until the time to reverse the current direction
            rest_voltage = time.time()-base_time - voltage_time
            while rest_voltage<0 and rest_until_target_time<0: # recird voltage vales, priotizing current flipping
                time.sleep(time_accuracy)
                rest_voltage = time.time()-base_time - voltage_time
                rest_until_target_time = time.time() - base_time - target_time
            sheet.cell(row = int(voltage_time/voltage_interval)+1, column = 12).value = time.time()-base_time
            sheet.cell(row = int(voltage_time/voltage_interval)+1, column = 13).value = keithley2.voltage # record to the excel file
            voltage_time = voltage_time + voltage_interval #reset the time to record voltage
            rest_until_target_time = time.time() - base_time - target_time
        current = current*(-1)
        target_time = target_time + half_period_time[i]
        rest_until_target_time = time.time() - base_time - target_time

    # at last, apply extra current for 1 sec to justify the calculation of Tox-Tred(最後に、1秒だけ余分に電流を流す。Tox-Tredの計算の帳尻を合わせるため。)
    keithley2.ramp_to_current(current,steps=3)
    time.sleep(1)
    keithley2.shutdown()
    sheet.cell(row = 11, column = 5).value = voltage_interval # you can save voltage_interval outside of the function
    sheet.cell(row = 6, column = 5).value = current*1000 # record current to the excel file (current was difined in this function)
                                                        
    print('current_apply_done.')

def calculation():
    global half_period_points
    starting_cell = 2 # row number of the cell corresponding to 0 s
#    starting_cell_calc = starting_cell + half_period_points*2 # row number of the cell corresponding to 0 s
#    print("test, starting_cell:", starting_cell_calc)

    ### average temperature of each points within the cycle###
    sheet.cell(row = 1, column = 7).value = 't (s)'
    for a in range(0, half_period_points * 2):
        sheet.cell(row = starting_cell + a, column = 7).value = a * pulse_interval
    sheet.cell(row = 1, column = 8).value = 'Average temperature (C)'
    for b in range(0,half_period_points * 2): 
        total = 0
        row_number = 0
        for step in range (0, period_number[i]):
            total = total + (sheet.cell(row = (starting_cell + half_period_points * 2) + row_number + b, column = 2).value) # summation 
            #1st cycle is rejected
            row_number = row_number + half_period_points * 2 # hop n to n+1, by using row_number
        #print((b,",","total:",total))
        sheet.cell(row = starting_cell + b, column = 8).value = total/period_number[i] #averaged

    ### time of 1 cycle ###
    sheet.cell(row = 1, column = 9).value = 't (s)'
    for time_monitored in range(0, half_period_points):
        sheet.cell(row = starting_cell+a, column = 9).value = time_monitored * pulse_interval 

    ### Tox(t) - Tred(t) ### Assuming oxidative current applied first to the electrode of the thermistor side 
    sheet.cell(row = 1, column = 10).value = 'Tox(t)-Tred(t) (K)'
    half_period_points = int(half_period_time[i]/pulse_interval) #obtain needed points for calculation
    for c in range(0, half_period_points):
        sheet.cell(row = starting_cell + c, column = 10).value = sheet.cell(row = starting_cell + c, column = 8).value - sheet.cell(row = starting_cell + c + half_period_points, column = 8).value 

    print('cauculation_done.')
  
""" # You can create charts on the excel file, if you want. (But this function tends to raise errors)
def chart_creation():
    chart_1 = ScatterChart() # Somehow creating a large chart and filling makes an error, thus charts were devided　なぜか1つのchartに代入しなおす方法にするとエラーを吐くため、3つのchartを分けている.
    chart_2 = ScatterChart() 
    chart_3 = ScatterChart()
    chart_4 = ScatterChart()
 
    start_row = 2 # row (all graph have the same row)  
    # raw data
    end_row_1 = (5* half_period_points*2) +(start_row-1) # column 列。5周期だけプロットする
    x_1 = Reference(sheet, min_col = 1, min_row = start_row , max_row = end_row_1) # sheetの1列目start_row行目からend_row_1行目までのデータを取得。
    y_1 = Reference(sheet, min_col = 2, min_row = start_row , max_row = end_row_1) # sheetの2行目start_row行目からend_row_1行目までのデータを取得。
    series_1 = Series(y_1, x_1, title = "Temperature vs. time")
    chart_1.legend = None
    chart_1.series.append(series_1) # seriesオブジェクトをもとにグラフ作成
    chart_1.x_axis.title = "time"
    chart_1.y_axis.title = "Temperature"
    sheet.add_chart(chart_1,"A7")#　plot 
    
    # Average temperature
    end_row_2 = half_period_points*2 +(start_row-1)
    x_2 = Reference(sheet, min_col = 7, min_row = start_row , max_row = end_row_2) 
    y_2 = Reference(sheet, min_col = 8, min_row = start_row , max_row = end_row_2) 
    series_2 = Series(y_2, x_2, title = "Average temperature vs. time") 
    chart_2.legend = None
    chart_2.x_axis.title = "time"
    chart_2.y_axis.title = "Average temperature"
    chart_2.series.append(series_2)
    sheet.add_chart(chart_2, "I7")

    # Tox-Tred 
    end_row_3 = half_period_points+(start_row-1)
    x_3 = Reference(sheet, min_col = 9, min_row = start_row , max_row = end_row_3) 
    y_3 = Reference(sheet, min_col = 10, min_row = start_row , max_row = end_row_3) 
    series_3 = Series(y_3, x_3, title = "Tox(t)-Tred(t) vs. t")
    chart_3.legend = None
    chart_3.series.append(series_3)
    chart_3.x_axis.title = "t"
    chart_3.y_axis.title = "Tox(t)-Tred(t)"
    sheet.add_chart(chart_3, "Q7")
    
    # voltage
    end_row_4 = 5* half_period_points*2 +(start_row-1) 
    x_4 = Reference(sheet, min_col = 12, min_row = start_row , max_row = end_row_4) 
    y_4 = Reference(sheet, min_col = 13, min_row = start_row , max_row = end_row_4) 
    series_4 = Series(y_4, x_4, title = "Voltage vs. time")
    chart_4.legend = None
    chart_4.series.append(series_4)
    chart_4.x_axis.title = "time"
    chart_4.y_axis.title = "Voltage"
    sheet.add_chart(chart_4,"Y7")

    print('chart_creation done.')
"""

""" # can send you a slack message after the completion of the measurement, if you want
def slack_notify(msg = 'measurement finished'):
    slack_user_id = 'Type your slack id here'
    slack_webhook_url = 'Type your slack URL here'
    requests.post(slack_webhook_url, json={"text":f"<@{slack_user_id}> {msg}"}) 
    requests.post(slack_webhook_url, json={"text":f" {msg}"})
"""


### conducting functions ###

if __name__ == "__main__":
    sample_name = 'Put your sample name here'
    # You can carry out 4 measurements in a row. (Actually, len(I), len(period_number) or len(half_period_time). You can change the length of the list if you want.)
    I = np.array([0.1,0.2,0.3,0.4]) # input the amplitude of current [mA]
    period_number = [5000,5000,5000,5000] #How many cycles in measurements. 
    half_period_time = [4,4,4,4] # The timing of current flip [s]
    before_measurement_h = 8 # waiting time for thermal equilibrium, at least 3 [hour], (this is the empirical value, you don't have to obey)
    measurement_interval_h = 2 # interval waiting time between measurements [hour]   

    measurements = int(input('number of measrurements: ')) # the number of the measurements

    # ユーザ名メモ。Peltier, kimilb2019_8, fmato
    # C:\Users\kimilab2019_8\Desktop\filename.xlsx

    time_calculation() # estimated time (actual measurement time is a bit longer)
    ec_confirmation()
    confirmation=int(input('conduct measuremant : 1')) # Put 1, when experimental conditions are alright, . Otherwise measurement won't be executed

    np.array(I, dtype=float)# convert to float
    I = I/1000 # convert from [mA] to [A]
    np.array(period_number, dtype=int) # convert to integer
    np.array(half_period_time, dtype=float)# convert to float

    if confirmation == 1: # when condition is judged alright
        connection_test()
        print(datetime.now())
        time.sleep(before_measurement_h*3600)
        day = datetime.now().strftime("%Y_%m_%d")
        path_peltier = "Put your path here"
        path = path_peltier + day
        if os.path.exists(path) == False:
            os.chdir(path_peltier)
            os.mkdir(day)
        else:
            pass
        for i in range(measurements):
            print(datetime.now())
            step = 0
            book = openpyxl.Workbook() # make a excel file
            sheet = book.worksheets[0] 
            initial_settings() #turn on sourcemeters
            half_period_points = int(half_period_time[i]/pulse_interval) 
            required_time = (period_number[i]+1) * half_period_time[i] *2 #calculate measurement time [s]
            thermistor_steps = int(required_time/pulse_interval)+int(1/pulse_interval) 

            sheet.cell(row = 7, column = 5).value = period_number[i]+1
            sheet.cell(row = 8, column = 5).value = half_period_time[i]

            base_time = time.time()
            executor = concurrent.futures.ThreadPoolExecutor(max_workers=2) # perform in multithreads
            executor.submit(thermistor)
            executor.submit(current_apply(I[i]))
            executor.shutdown()
            date = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
            book.save("{0}/{2}_{3}mA_{4}s_{5}+1_{1}.xlsx".format(path, date, sample_name, I[i]*1000, half_period_time[i], period_number[i]))
            calculation()
            #chart_creation()
            book.save("{0}/{2}_{3}mA_{4}s_{5}+1_{1}.xlsx".format(path, date, sample_name, I[i]*1000, half_period_time[i], period_number[i]))
            if i == measurements-1:
                break
            time.sleep(measurement_interval_h * 3600)
        print('finished.')

    else: 
        print('cancelled')
